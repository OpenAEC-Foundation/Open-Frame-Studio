"""Kozijnstaat (window schedule) generation in PDF and Excel formats."""

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


PANEL_LABELS = {
    "fixed_glass": "Vast glas",
    "turn_tilt": "Draaikiepraam",
    "turn": "Draairaam",
    "tilt": "Kiepraam",
    "sliding": "Schuifraam",
    "door": "Deur",
    "panel": "Paneel",
    "ventilation": "Ventilatie",
}

# OpenAEC brand colors
AMBER = colors.HexColor("#D97706")
DEEP_FORGE = colors.HexColor("#36363E")
WARM_GOLD = colors.HexColor("#F59E0B")


def generate_kozijnstaat(project_data: dict, output_path: str, fmt: str = "pdf"):
    """Generate a kozijnstaat (window schedule) from project data."""
    if fmt == "xlsx":
        _generate_xlsx(project_data, output_path)
    else:
        _generate_pdf(project_data, output_path)


def _generate_pdf(project_data: dict, output_path: str):
    """Generate PDF kozijnstaat."""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "KozijnstaatTitle",
        parent=styles["Title"],
        fontSize=18,
        textColor=DEEP_FORGE,
        spaceAfter=6 * mm,
    )
    subtitle_style = ParagraphStyle(
        "KozijnstaatSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.gray,
        spaceAfter=10 * mm,
    )

    elements = []

    # Title
    project_info = project_data.get("projectInfo", {})
    title_text = f"Kozijnstaat — {project_info.get('name', 'Project')}"
    elements.append(Paragraph(title_text, title_style))
    elements.append(Paragraph(
        f"Projectnummer: {project_info.get('number', '-')} | "
        f"Opdrachtgever: {project_info.get('client', '-')}",
        subtitle_style,
    ))

    # Table data
    headers = [
        "Merk", "Naam", "Breedte\n(mm)", "Hoogte\n(mm)",
        "Materiaal", "Kolommen", "Rijen", "Cellen",
        "Paneel types", "Beglazing", "Kleur binnen", "Kleur buiten",
    ]

    table_data = [headers]

    for kozijn in project_data.get("kozijnen", []):
        frame = kozijn.get("frame", {})
        cells = kozijn.get("cells", [])
        grid = kozijn.get("grid", {})

        # Cell type summary
        type_counts = {}
        for cell in cells:
            pt = cell.get("panelType", "fixed_glass")
            label = PANEL_LABELS.get(pt, pt)
            type_counts[label] = type_counts.get(label, 0) + 1
        type_summary = ", ".join(f"{v}x {k}" for k, v in type_counts.items())

        # Material label
        material = frame.get("material", {})
        if isinstance(material, dict):
            mat_key = list(material.keys())[0] if material else "onbekend"
            mat_val = list(material.values())[0] if material else ""
            mat_label = f"{mat_key} ({mat_val})" if mat_val else mat_key
        else:
            mat_label = str(material)

        # Glazing from first cell
        glazing = cells[0].get("glazing", {}) if cells else {}
        glaz_label = f"{glazing.get('glassType', '-')} {glazing.get('thicknessMm', '')}mm"

        row = [
            kozijn.get("mark", "-"),
            kozijn.get("name", "-"),
            str(int(frame.get("outerWidth", 0))),
            str(int(frame.get("outerHeight", 0))),
            mat_label,
            str(len(grid.get("columns", []))),
            str(len(grid.get("rows", []))),
            str(len(cells)),
            type_summary,
            glaz_label,
            frame.get("colorInside", "-"),
            frame.get("colorOutside", "-"),
        ]
        table_data.append(row)

    # Create table
    col_widths = [35*mm, 45*mm, 25*mm, 25*mm, 35*mm, 20*mm, 20*mm, 20*mm, 50*mm, 35*mm, 30*mm, 30*mm]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), DEEP_FORGE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        # Body
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (2, 1), (7, -1), "CENTER"),
        ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("LINEBELOW", (0, 0), (-1, 0), 2, AMBER),
        # Alternating rows
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F4")]),
        # Padding
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 10 * mm))

    # Footer
    footer_style = ParagraphStyle(
        "Footer",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.gray,
    )
    elements.append(Paragraph(
        "Gegenereerd door Open Frame Studio — OpenAEC Foundation",
        footer_style,
    ))

    doc.build(elements)


def _generate_xlsx(project_data: dict, output_path: str):
    """Generate Excel kozijnstaat."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kozijnstaat"

    # Styles
    header_font = Font(name="Inter", bold=True, color="FFFFFF", size=9)
    header_fill = PatternFill(start_color="36363E", end_color="36363E", fill_type="solid")
    amber_border = Border(bottom=Side(style="medium", color="D97706"))
    body_font = Font(name="Inter", size=9)
    center_align = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "Merk", "Naam", "Breedte (mm)", "Hoogte (mm)",
        "Materiaal", "Kolommen", "Rijen", "Cellen",
        "Paneel types", "Beglazing", "Kleur binnen", "Kleur buiten",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = amber_border

    # Data rows
    for row_idx, kozijn in enumerate(project_data.get("kozijnen", []), 2):
        frame = kozijn.get("frame", {})
        cells_data = kozijn.get("cells", [])
        grid = kozijn.get("grid", {})

        type_counts = {}
        for c in cells_data:
            pt = c.get("panelType", "fixed_glass")
            label = PANEL_LABELS.get(pt, pt)
            type_counts[label] = type_counts.get(label, 0) + 1
        type_summary = ", ".join(f"{v}x {k}" for k, v in type_counts.items())

        material = frame.get("material", {})
        if isinstance(material, dict):
            mat_label = str(list(material.keys())[0]) if material else "onbekend"
        else:
            mat_label = str(material)

        glazing = cells_data[0].get("glazing", {}) if cells_data else {}

        values = [
            kozijn.get("mark", "-"),
            kozijn.get("name", "-"),
            int(frame.get("outerWidth", 0)),
            int(frame.get("outerHeight", 0)),
            mat_label,
            len(grid.get("columns", [])),
            len(grid.get("rows", [])),
            len(cells_data),
            type_summary,
            f"{glazing.get('glassType', '-')} {glazing.get('thicknessMm', '')}mm",
            frame.get("colorInside", "-"),
            frame.get("colorOutside", "-"),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            if 3 <= col_idx <= 8:
                cell.alignment = center_align

    # Auto-fit column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

    wb.save(output_path)
